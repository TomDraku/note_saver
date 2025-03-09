import openpyxl


txt_file_path = "notatka_09_03.txt"
excel_file_path = "Exel_notatka.xlsx"


# opaning a note file txt
with open(txt_file_path, 'r', encoding='utf-8') as txt_file:
    content = txt_file.read()

    paragraphs = content.split('\n\n')  # Separation of paragraphs
   
   
    # creating an object 
    workbook = openpyxl.Workbook()
    sheet = workbook.active # creating a sheet

    for index, paragraph in enumerate(paragraphs):
            for index, paragraph in enumerate(paragraphs):
                try:
                    parts = paragraph.strip().split('-')
                    sheet.cell(row=index + 1, column=2, value=parts[0]) # save in column B
                    sheet.cell(row=index + 1, column=3, value=parts[1]) # save in column C
                except IndexError:
                    print(f"Error: Missing '-' character in paragraph: {paragraph}")

        
    workbook.save(excel_file_path)
    print(f"Date was saves in Excel file: {excel_file_path}")